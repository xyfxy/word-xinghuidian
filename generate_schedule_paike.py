#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一键排课项目实施计划排期表生成器
2025年9月-11月15日工作排期
"""

import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_schedule():
    """创建一键排课项目实施计划排期表"""
    
    # 项目基本信息
    project_info = {
        '项目名称': '一键排课项目实施计划',
        '项目周期': '2025年9月1日-11月15日（约11周，考虑国庆假期）',
        '项目目标': '开发并部署一键排课系统，实现智能化课程安排',
        '技术架构': '前后端分离架构 + AI算法 + 数据库系统',
        '工作安排': '需求分析→系统设计→开发实施→测试部署→培训交付',
        '假期安排': '国庆节10月1日-7日放假，周末正常休息'
    }
    
    # 2025年9月-11月日历分析
    # 9月1日（星期一）开始
    # 国庆节：10月1日-10月7日（7天假期）
    # 项目结束：11月15日（星期六）
    
    # 详细任务排期数据（基于schedule.md内容，压缩为11周）
    schedule_data = [
        # 第一阶段：基础架构与核心数据引擎（保留调研）
        ['第一阶段\n基础架构与核心数据引擎', '第1周', '9月1日-9月5日\n（星期一至五）', 
         '• 项目启动与数据基础调研\n• 定义并搭建"智能资源中心"数据库\n• 开发初版数据同步脚本，实现从OA系统到本地数据库的一次性数据迁移', 
         '• 项目代码骨架\n• 可运行的数据库实例\n• 能够拉取全量资源数据的同步脚本', 
         '开发团队', '项目启动与调研'],
         
        ['', '第2周', '9月8日-9月12日\n（星期一至五）',
         '• RAG引擎构建：数据处理与向量化\n• 为资源（教师/现场教学）设计并实现文本预处理流程\n• 集成向量嵌入模型，开发向量化脚本\n• 将所有资源处理并存入向量数据库(Vector Store)',
         '• 一套完整的数据处理与向量化流程\n• 包含所有资源的、可被检索的向量索引库',
         '开发团队', 'RAG引擎构建'],
         
        ['', '第3周', '9月15日-9月19日\n（星期一至五）',
         '• 节点一开发：需求解析\n• 设计策略指令集的JSON结构\n• 开发需求表单的解析逻辑\n• 集成Web acp搜索工具，实现客户背景（政府/国企）的自动化分析\n• 节点一完善与检索测试',
         '• "节点一"模块雏形，能将原始需求转化为结构化的策略指令集\n• 一个经过测试、功能完备的"节点一"\n• 初步的检索质量评估报告',
         '开发团队', '节点一开发完善'],
         
        # 第二阶段：智能匹配与方案草案生成
        ['第二阶段\n智能匹配与方案草案生成', '第4周', '9月22日-9月26日\n（星期一至五）',
         '• 节点二开发：硬性过滤与语义检索\n• 开发资源硬性过滤器模块（地点、人数、开发状态等）\n• 将硬性过滤与语义检索结合，形成两步筛选流程\n• 输出满足硬性条件并按主题相关度排序的候选资源列表',
         '• 能返回符合硬性约束的相关候选资源的模块',
         '开发团队', '节点二开发'],
         
        ['', '第5周', '9月29日-9月30日\n（星期一二）',
         '• 节点二完善：专家经验评分模型\n• 设计包含完整评分维度的资源数据结构\n• 实现多维度专家评分模型（五有、次新、多样性、模板、反馈）',
         '• 功能完整的"节点二"，能输出包含综合得分和推荐理由的候选资源列表',
         '开发团队', '国庆前完成，10月1-7日放假'],
         
        # 国庆假期
        ['国庆假期', '国庆', '10月1日-10月7日\n（国庆7天假期）',
         '• 国庆节放假休息\n• 团队休整充电\n• 准备节后开发工作',
         '• 假期休息\n• 准备节后工作计划',
         '全体人员', '国庆节法定假期'],
         
        ['', '第6周', '10月8日-10月10日\n（星期三至五）',
         '• 利用LLM生成高度相关的"一句话推荐理由"\n• 节点三开发：约束求解引擎\n• 设计日程编排的数据结构（日程表）',
         '• "节点三"雏形，能根据候选资源和核心约束，生成一个逻辑通顺的方案草案',
         '开发团队', '节后功能开发'],
         
        ['', '第7周', '10月13日-10月17日\n（星期一至五）',
         '• 实现核心硬性约束的校验逻辑（配比、预算、新资源数量等）\n• 开发一个基于贪心或启发式搜索的初步日程生成算法\n• 节点三完善：回溯与优化\n• 实现所有排程逻辑约束（如抵达/返程日规则、调研式教学位置）',
         '• 一个具备自我修复能力的日程生成引擎\n• 第二阶段评审',
         '开发团队', '约束求解完善'],
         
        # 第三阶段：方案优化、整合与部署
        ['第三阶段\n方案优化、整合与部署', '第8周', '10月20日-10月24日\n（星期一至五）',
         '• 引入软性约束作为优化目标（如主题连续性）\n• 开发回溯机制，当生成失败时，能自动尝试次优选项或向上游请求更多资源\n• 节点四开发：方案智能美化\n• 开发"节点四"模块，接收日程草案',
         '• 一个能将方案草案自动优化和包装成专业文档内容的"节点四"',
         '全栈团队', '方案优化'],
         
        ['', '第9周', '10月27日-10月31日\n（星期一至五）',
         '• 设计并实现针对不同美化任务（模块标题、课程点题、推荐理由重写）的LLM Prompts\n• 实现自动化审核逻辑，对照Checklist进行程序化校验\n• 端到端流程整合与测试\n• 将四个独立的智能节点串联成一个完整的自动化工作流',
         '• 一个可从原始需求一步生成到最终方案的完整系统\n• 端到端测试报告',
         '全栈团队', '系统集成'],
         
        ['', '第10周', '11月3日-11月7日\n（星期一至五）',
         '• 构建并管理在节点间流转的"全局方案状态"对象\n• 收集真实的OA需求作为测试集，进行完整的端到端流程测试\n• 系统调优与最终评估\n• 组织业务专家对测试结果进行人工评估',
         '• 一个经过验证和调优的、性能稳定的系统\n• 与下游系统对接的标准数据接口',
         '测试团队', '系统测试验证'],
         
        ['', '第11周', '11月10日-11月14日\n（星期一至五）',
         '• 根据反馈，精调评分权重、LLM Prompts和业务规则\n• 开发最终的格式化输出模块，确保数据能被OA系统无缝导入\n• 部署、文档与交付\n• 配置生产环境，并通过API接口部署整个系统\n• 编写详细的技术实现文档和给业务团队的用户操作手册',
         '• 生产环境中可用的"一键排课"系统v1.0\n• 完整的技术与用户文档',
         '运维+培训团队', '部署培训交付'],
         
        ['项目交付', '收尾', '11月15日（星期六）',
         '• 进行项目总结，正式交付v1.0版本\n• 客户验收签字，项目结项\n• 项目归档和经验分享',
         '• 项目正式交付\n• 验收文档签署\n• 项目归档完成',
         '全体团队', '项目圆满完成']
    ]
    
    # 创建DataFrame
    df = pd.DataFrame(schedule_data, columns=[
        '阶段', '时间', '日期', '核心任务', '交付成果', '负责人', '备注'
    ])
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    
    # 只创建项目排期表
    ws1 = wb.active
    ws1.title = "项目排期"
    
    # 设置样式
    header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
    content_font = Font(name='Microsoft YaHei', size=10)
    title_font = Font(name='Microsoft YaHei', size=14, bold=True)
    
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    stage_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    holiday_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 工作表1：项目排期主表
    ws1.title = "项目排期"
    
    # 添加项目信息标题
    ws1.merge_cells('A1:G1')
    ws1['A1'] = '一键排课项目实施计划排期表 - 2025年9月-11月'
    ws1['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True)
    ws1['A1'].alignment = center_alignment
    
    # 添加项目基本信息
    info_row = 3
    for key, value in project_info.items():
        ws1.merge_cells(f'A{info_row}:B{info_row}')
        ws1[f'A{info_row}'] = key
        ws1[f'A{info_row}'].font = Font(name='Microsoft YaHei', size=10, bold=True)
        ws1.merge_cells(f'C{info_row}:G{info_row}')
        ws1[f'C{info_row}'] = value
        ws1[f'C{info_row}'].font = content_font
        ws1[f'C{info_row}'].alignment = left_alignment
        info_row += 1
    
    # 添加表头
    headers = ['阶段', '时间', '日期', '核心任务', '交付成果', '负责人', '备注']
    header_row = info_row + 2
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # 添加数据
    for row_idx, row_data in enumerate(schedule_data, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = content_font
            cell.alignment = left_alignment if col_idx > 2 else center_alignment
            cell.border = thin_border
            
            # 阶段单元格特殊格式
            if col_idx == 1 and value and '阶段' in value:
                cell.fill = stage_fill
                cell.font = Font(name='Microsoft YaHei', size=10, bold=True)
            
            # 国庆假期特殊标记
            if col_idx == 1 and value == '国庆假期':
                cell.fill = holiday_fill
                cell.font = Font(name='Microsoft YaHei', size=10, bold=True)
    
    # 设置列宽
    column_widths = [15, 12, 18, 40, 35, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 设置行高
    for row in range(header_row + 1, header_row + len(schedule_data) + 1):
        ws1.row_dimensions[row].height = 80
    
    # 保存文件
    import time
    filename = f'Schedule_Paike_2025_Sep_Nov_{int(time.time())}.xlsx'
    wb.save(filename)
    
    print(f"排期表已生成: {filename}")
    print("\n项目排期表包含完整的任务安排（9月1日-11月15日）")
    
    return filename

if __name__ == "__main__":
    print("开始生成一键排课项目实施计划排期表（2025年9月-11月15日）...")
    filename = create_schedule()
    print(f"\n排期表生成完成！文件名：{filename}")
    print("\n工作安排：")
    print("- 第一阶段（9月1-14日）：需求分析与方案设计")
    print("- 第二阶段（9月15日-10月31日）：系统开发（国庆放假7天）")
    print("- 第三阶段（11月1-15日）：测试部署与交付")
    print("\n特别提醒：")
    print("- 国庆节10月1-7日放假，注意提前安排工作")
    print("- 周末正常休息，工作日全力推进")
    print("- 11月15日（周六）项目交付，需要加班一天")