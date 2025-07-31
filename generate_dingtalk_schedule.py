#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Word星辉点钉钉工作台部署排期表生成器
2025年8月工作排期 - 基于项目实际情况制定
"""

import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_schedule():
    """创建钉钉工作台部署排期表"""
    
    # 项目基本信息
    project_info = {
        '项目名称': 'Word星辉点智能编辑器钉钉工作台部署',
        '项目周期': '2025年9月1日-11月14日（11周）',
        '项目目标': '将Word星辉点项目部署到钉钉工作台，实现仅通过钉钉访问的安全控制',
        '技术架构': 'React18 + Node.js + Express + TypeScript + 钉钉H5微应用',
        '工作安排': '需求调研分析→功能实现（多阶段）→平台认知→架构设计→集成调试→UI/UX优化→培训交付（11周完整流程）',
        '当前状态': '基础功能开发完成，包含AI生成、模板系统、Word导入导出、MaxKB集成',
        '系统特点': '无需用户登录系统，不需要集成钉钉免登功能，仅需嵌入钉钉工作台',
        '开发环境': '钉钉企业内部应用支持HTTP，可使用localhost或内网穿透，无需SSL证书'
    }
    
    # 9月2025年日历分析：9月1日（星期一）开始
    # 11周完整排期：包含需求调研到最终交付的完整流程
    
    # 详细任务排期数据（基于图片内容）
    schedule_data = [
        # 第1周：需求调研分析（9月第一周）
        ['第一周\n需求调研分析', '9月1-5日', '9月1日-9月5日\n（星期一至五）', 
         '• 全功能模块「智能文档」的调研\n• 功能规划和项目指标（文本生成和知识库集成等）方面梳理\n• 建构网站系统框架（前端：React，后端：Express）', 
         '可运行的前端页面框架\n\n能在前端页面通过MaxKb进行搜索\n\n实现前端和后端的网络请求框架', 
         '项目团队', '需求调研，了解真实需求'],
         
        # 第2周：功能实现-数据库连接（9月第二周）
        ['第二周\n功能实现1', '9月8-12日', '9月8日-9月12日\n（星期一至五）',
         '• 功能素：编辑（保存）：设计全面填充文本和AI填充模板。\n• 基础编辑功能构建：添加组块绘画。\n• 自制组块系统和界面大小调整系统（Render Resize）',
         '• 素以填充（插入）功能代替应用数据集储存文件；\n• 包含新的「填充」功能数据统合包。',
         '开发团队', '核心功能开发'],
         
        # 第3周：平台认知+页面部署（9月第三周）
        ['第三周\n平台认知+页面部署', '9月15-19日', '9月15日-9月19日\n（星期一至五）',
         '• 以对模板功能深入认识和修改。\n• 设置内容学习和模板的输入。\n• Webcom web接口文档（网页，调用服务，后台接口 MaxKb）API',
         '• 平台一期认知UI。全功能交互系统完成和认知支付功能数据总集',
         '前端开发团队', '界面优化和集成'],
         
        # 第4周：功能实现2（9月第四周）
        ['第四周\n功能实现2', '9月22-26日', '9月22日-9月26日\n（星期一至五）',
         '• 企业规则模板功能（宏观，最细分等级）的CRUD。\n• 解决内容保存，临床疗程成果分析展示。\n• 解析平台一主视图机理模式，HTML创建多种不同预定项模块在复杂任务中保准',
         '对项目保留详细设置和架构。\n\n第一阶段升级',
         '后端开发团队', '功能完善'],
         
        # 第5周：架构设计与应用优化（9月29-30日）
        ['第五周\n架构设计与应用优化', '9月29-30日', '9月29日-9月30日\n（星期一至二）',
         '• 有关新增能反应反应模板设计（通过一手数据和工程中相应功能）\n• 机器接口最完整后端（监测网云表最复杂，部署进化调试和信息等）\n• 配置分配并行关系（通过数据动态定立）接口分层标准规范（如端模调整增强项模块）',
         '• 展现识别现有的设计方式现有数据架构等表（支持创建多种）的重构综合报告，\n• 分表方法最终应认定数据库架构和数据并行化系统',
         '架构团队', '架构优化和设计'],
         
        # 第6周：集成调试+在线部署（10月第一周）
        ['第六周\n集成调试+在线部署', '10月6-10日', '10月6日-10月10日\n（星期一至五）',
         '• 对目标功能和实施流程打包策略（启用，小装，访问目录，蓝板，启用）\n• 动态123支持现流程预构架-——功能地球联络。',
         '对遗数据库（系统，目标）并且现有多标准不需升级（不需要进行原接触功能分离）\n• 如果还面临大幅度表示集和事件现实，将可以作为自由的数据集提供控制',
         '运维团队', '系统部署和优化'],
         
        # 第7周：功能实现3-导出和生成（10月第二周）
        ['第七周\n功能实现3', '10月13-17日', '10月13日-10月17日\n（星期一至五）',
         '• 分页输出编程最适量服务（AI测层）\n• 全增的生成机架建设原版应用，仅对于建设建和文件。\n• 完全一体下少出式系统式样管理改分网络自主建系统。',
         '• 目标实现，智能监视组拿到目标文件上。\n• 综合应用设计系统',
         '开发团队', '导出功能完善'],
         
        # 第8周：功能实现4-知识库集成（10月第三周）
        ['第八周\n功能实现4', '10月20-24日', '10月20日-10月24日\n（星期一至五）',
         '• 实现最后建设管理和维护功能测试要求（已经完成），解析对策文件。\n• 开发AI仓库（前端端设计）基于多服务更系统。\n• 分析多个整数计划（建统类，智能保护和交互设建议和设计完成）',
         '• 功能整合，应用，效率和进度应用。\n• 结算上线升级功能',
         '全栈开发团队', '知识库功能集成'],
         
        # 第9周：UI/UX设计-页面优化（10月第四周）
        ['第九周\nUI/UX设计优化', '10月27-31日', '10月27日-10月31日\n（星期一至五）',
         '• 设计用户体验并且编写产权更让人大认识。\n• 提请全新，精准面的设计（UX Elements如版本面）。\n• 分导解性能并设置用更好，更独具数服用和完成和用户总结。',
         '5个主要系统能清体质高度推数据。\n• 分享能系统的应对总的规划（暂新）',
         'UI/UX设计团队', '用户体验优化'],
         
        # 第10周：原码应用第三周培训（11月第一周）  
        ['第十周\n原码应用第三周培训', '11月3-7日', '11月3日-11月7日\n（星期一至五）',
         '• 配置分别标配，示例定义和对应的交叉需复要素文件。\n• 展示源数据变更文件对Scripting功能创建分的时间总体数量基。\n• 分叉可提取制度，企业版用户，调度表。',
         '可展现设计大体代代服务。\n• 评阶代表评评',
         '培训团队', '系统培训和交付'],
         
        # 第11周：验收测试和正式交付（11月第二周）
        ['第十一周\n验收测试和交付', '11月10-14日', '11月10日-11月14日\n（星期一至五）',
         '• 对用户和完整制造测试数据（已完整数据路线上文分析）。\n• 创建应用与试运行ADS设计全面用程指标，应用\n• 主要模式可以通过多个数据测试。对设置对项明原后模块合作对应',
         '• 项目验收通过并正式对接系统地点和平稳\n• 模板综合项目评价',
         '项目团队+客户', '最终验收交付']
    ]
    
    # 创建DataFrame
    df = pd.DataFrame(schedule_data, columns=[
        '阶段', '周次', '日期', '核心任务', '交付成果', '负责人', '备注'
    ])
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认工作表并创建新工作表
    wb.remove(wb.active)
    ws1 = wb.create_sheet("项目排期", 0)
    ws2 = wb.create_sheet("风险控制", 1)
    ws3 = wb.create_sheet("资源配置", 2)
    ws4 = wb.create_sheet("技术方案", 3)
    
    # 设置样式
    header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
    content_font = Font(name='Microsoft YaHei', size=10)
    title_font = Font(name='Microsoft YaHei', size=14, bold=True)
    
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    stage_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 工作表1：项目排期主表
    ws1.title = "项目排期"
    
    # 添加项目信息标题
    ws1.merge_cells('A1:G1')
    ws1['A1'] = 'Word星辉点钉钉工作台部署排期表 - 2025年9月至11月（完整版：11周流程）'
    ws1['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True)
    ws1['A1'].alignment = center_alignment
    
    # 添加项目基本信息
    info_row = 3
    for key, value in project_info.items():
        ws1.merge_cells(f'A{info_row}:B{info_row}')
        ws1[f'A{info_row}'] = key
        ws1[f'A{info_row}'].font = Font(name='Microsoft YaHei', size=10, bold=True)
        ws1.merge_cells(f'C{info_row}:G{info_row}')
        ws1[f'C{info_row}'] = value
        ws1[f'C{info_row}'].font = content_font
        ws1[f'C{info_row}'].alignment = left_alignment
        info_row += 1
    
    # 添加表头
    headers = ['阶段', '周次', '日期', '核心任务', '交付成果', '负责人', '备注']
    header_row = info_row + 2
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # 添加数据
    for row_idx, row_data in enumerate(schedule_data, header_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = content_font
            cell.alignment = left_alignment if col_idx > 2 else center_alignment
            cell.border = thin_border
            
            # 阶段单元格特殊格式
            if col_idx == 1 and value and '阶段' in value:
                cell.fill = stage_fill
                cell.font = Font(name='Microsoft YaHei', size=10, bold=True)
    
    # 设置列宽
    column_widths = [15, 8, 12, 35, 30, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 设置行高
    for row in range(header_row + 1, header_row + len(schedule_data) + 1):
        ws1.row_dimensions[row].height = 80
    
    # 工作表2：风险控制（优化版）
    ws2.title = "风险控制"
    risk_data = [
        ['风险类型', '风险描述', '影响程度', '应对措施', '责任人', '监控指标'],
        ['技术风险', '钉钉嵌入兼容性问题', '低', '钉钉H5微应用技术成熟稳定；无需复杂API集成；风险很低', '技术负责人', '嵌入访问成功率>99%'],
        ['安全风险', '访问控制被绕过', '中', '前端检测钉钉环境+后端验证User-Agent+访问来源控制', '安全负责人', '非钉钉访问检测为0'],
        ['网络风险', '内网穿透稳定性问题', '低', '使用钉钉官方内网穿透工具；准备localhost备用方案', '运维负责人', '网络连接稳定率>99%'],
        ['性能风险', 'AI生成功能在钉钉环境响应慢', '低', '钉钉容器性能良好；继续优化AI调用和缓存策略', '后端负责人', '平均响应时间<3秒'],
        ['兼容性风险', '钉钉容器环境兼容问题', '低', '钉钉H5容器标准化程度高；重点测试移动端和PC端', '前端负责人', '主流版本兼容率100%'],
        ['时间风险', '5周内无法完成所有优化', '低', '合理分配5周时间，充分优化各模块；预留缓冲时间', '项目经理', '按时完成率>95%'],
        ['依赖风险', 'MaxKB服务不稳定', '中', '增加重试机制；准备降级方案；与MaxKB团队建立沟通机制', '技术负责人', 'MaxKB可用率>99.5%'],
        ['用户体验风险', '钉钉环境下用户体验不佳', '中', '充分的用户测试和反馈收集；快速响应用户问题', '产品负责人', '用户满意度>90%']
    ]
    
    # 添加风险控制表头和数据
    for row_idx, row_data in enumerate(risk_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.font = content_font
                cell.alignment = left_alignment
            cell.border = thin_border
    
    # 设置风险控制表列宽
    risk_column_widths = [12, 25, 10, 40, 12, 20]
    for col, width in enumerate(risk_column_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 工作表3：资源配置
    ws3.title = "资源配置"
    resource_data = [
        ['资源类型', '具体配置', '数量', '工作量', '关键技能', '备注'],
        ['人力资源', '前端开发工程师', '2人', '全职4周', 'React + TypeScript + 钉钉H5开发', '负责界面适配和钉钉集成'],
        ['', '后端开发工程师', '2人', '全职4周', 'Node.js + Express + 钉钉API', '负责后端接口和安全控制'],
        ['', '测试工程师', '1人', '全职2周', '钉钉应用测试 + 安全测试', '负责功能和安全测试'],
        ['', '运维工程师', '1人', '兼职4周', 'Docker + Nginx + 钉钉部署', '负责环境配置和上线部署'],
        ['', '产品经理', '1人', '兼职4周', '钉钉工作台产品经验', '负责需求确认和用户培训'],
        ['', '安全专家', '1人', '兼职1周', '企业应用安全 + 钉钉安全', '负责安全方案设计和审核'],
        ['技术资源', '钉钉企业账号', '1个', '1个月', '企业级权限', '需要管理员权限配合'],
        ['', '开发测试环境', '2套', '1个月', 'Docker容器化部署', '开发环境 + 预生产环境'],
        ['', '生产服务器', '1套', '长期', '高可用配置', '支持钉钉工作台访问的生产环境'],
        ['', 'SSL证书', '1个', '长期', 'HTTPS加密', '钉钉要求必须HTTPS访问'],
        ['外部依赖', '千问AI服务', '按需', '长期', 'API调用额度', '确保AI生成功能稳定'],
        ['', 'MaxKB服务', '按需', '长期', '知识库访问', '确保知识库功能正常'],
        ['', '钉钉技术支持', '按需', '项目期间', '钉钉官方技术咨询', '遇到问题时的技术支持']
    ]
    
    # 添加资源配置数据
    for row_idx, row_data in enumerate(resource_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.font = content_font
                cell.alignment = left_alignment
            cell.border = thin_border
    
    # 设置资源配置表列宽
    resource_column_widths = [12, 20, 8, 12, 25, 25]
    for col, width in enumerate(resource_column_widths, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 工作表4：技术方案
    ws4.title = "技术方案"
    
    # 技术方案详细说明
    tech_content = [
        ['钉钉集成技术方案', ''],
        ['1. 应用类型选择', 'H5微应用 - 将现有Web应用嵌入钉钉工作台'],
        ['2. 身份认证方案', '钉钉免登 - 通过钉钉JSAPI获取用户身份，移除原有登录系统'],
        ['3. 安全访问控制', '多重验证：域名白名单 + Referer检查 + 钉钉Token验证 + iframe防护'],
        ['4. 界面适配方案', '响应式设计适配钉钉容器，优化移动端体验'],
        ['5. API集成方案', '集成钉钉JSAPI：免登、分享、通知、文件上传等'],
        ['', ''],
        ['部署架构方案', ''],
        ['1. 现有架构保持', 'React前端 + Node.js后端 + Express API'],
        ['2. 新增钉钉适配层', '前端：钉钉JSAPI集成模块；后端：钉钉身份验证中间件'],
        ['3. 安全控制层', 'Nginx反向代理 + 域名白名单 + 访问控制策略'],
        ['4. 监控告警', '应用性能监控 + 异常告警 + 访问日志分析'],
        ['', ''],
        ['关键技术要点', ''],
        ['1. 钉钉免登实现', 'dd.ready -> dd.runtime.permission.requestAuthCode -> 后端验证'],
        ['2. 安全访问控制', 'if (referer !== dingtalk && !dingtalkToken) { return 403; }'],
        ['3. 移动端适配', 'viewport设置 + rem适配 + touch事件优化'],
        ['4. 性能优化', '代码分割 + 懒加载 + CDN加速 + 缓存策略'],
        ['', ''],
        ['上线部署流程', ''],
        ['1. 钉钉应用创建', '管理员在钉钉开放平台创建H5微应用'],
        ['2. 域名配置', '添加应用域名到钉钉白名单'],
        ['3. 权限申请', '申请所需的JSAPI权限'],
        ['4. 应用发布', '设置应用图标、描述、可见范围'],
        ['5. 工作台配置', '将应用添加到企业工作台'],
    ]
    
    # 添加技术方案内容
    for row_idx, (key, value) in enumerate(tech_content, 1):
        ws4.cell(row=row_idx, column=1, value=key).font = Font(name='Microsoft YaHei', size=11, bold=True) if value == '' else content_font
        ws4.cell(row=row_idx, column=2, value=value).font = content_font
        ws4.cell(row=row_idx, column=2).alignment = left_alignment
    
    # 设置技术方案表列宽
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 60
    
    # 保存文件
    import time
    filename = f'DingTalk_Schedule_2025_Sep_to_Nov_{int(time.time())}.xlsx'
    wb.save(filename)
    
    print(f"排期表已生成: {filename}")
    print("\n包含以下工作表：")
    print("1. 项目排期 - 详细的11周开发计划")
    print("2. 风险控制 - 7个主要风险的应对措施")
    print("3. 资源配置 - 人力和技术资源需求")
    print("4. 技术方案 - 钉钉集成的技术实现细节")
    
    return filename

if __name__ == "__main__":
    print("开始生成Word星辉点钉钉工作台部署排期表（2025年9月至11月）...")
    filename = create_schedule()
    print(f"\n排期表生成完成！文件名：{filename}")
    print("\n工作安排（11周完整版）：")
    print("- 第1周（9月1-5日）：需求调研分析")
    print("- 第2周（9月8-12日）：功能实现1-数据库连接")
    print("- 第3周（9月15-19日）：平台认知+页面部署")
    print("- 第4周（9月22-26日）：功能实现2")
    print("- 第5周（9月29-30日）：架构设计与应用优化")
    print("- 第6周（10月6-10日）：集成调试+在线部署")
    print("- 第7周（10月13-17日）：功能实现3-导出和生成")
    print("- 第8周（10月20-24日）：功能实现4-知识库集成")
    print("- 第9周（10月27-31日）：UI/UX设计优化")
    print("- 第10周（11月3-7日）：原码应用第三周培训")
    print("- 第11周（11月10-14日）：验收测试和交付")
    print("\n重要简化：系统无需登录功能，不需要集成钉钉免登，仅需嵌入工作台")
    print("\n安全保障：确保只能通过钉钉工作台访问，其他方式完全禁用")