import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

def create_clean_implementation_plan():
    # 创建项目数据
    data = {
        '阶段': [],
        '周次': [],
        '核心任务': [],
        '交付成果': []
    }
    
    # 第一阶段：系统安全与AI模型集成（2周）
    phase1_tasks = [
        {
            '周次': '第1周',
            '核心任务': [
                '系统安全加固',
                '• API安全机制实现',
                '• 内容审核和敏感词过滤'
            ],
            '交付成果': [
                '• 安全防护机制生效'
            ]
        },
        {
            '周次': '第2周',
            '核心任务': [
                'AI模型厂商模块',
                '• 多AI厂商接入架构设计',
                '• 集成文心一言、通义千问、智谱清言',
                '• AI模型管理界面开发',
                '• 统一调用接口实现'
            ],
            '交付成果': [
                '• 支持3+AI模型切换',
                '• AI配置界面完成',
                '• 接口调用稳定'
            ]
        }
    ]
    
    # 第二阶段：核心功能开发（2周）
    phase2_tasks = [
        {
            '周次': '第3周',
            '核心任务': [
                'Word表格解析功能',
                '• 表格识别和解析引擎开发',
                '• 复杂表格处理（合并单元格、嵌套）',
                '• 表格编辑器组件',
                '• 导出格式保真优化'
            ],
            '交付成果': [
                '• 表格解析功能完成',
                '• 支持主流表格格式',
                '• 编辑和导出正常'
            ]
        },
        {
            '周次': '第4周',
            '核心任务': [
                'MaxKB工作流集成',
                '• 两个培训报告模板开发',
                '• 固定内容与AI内容混合排版',
                '• MaxKB知识库配置',
                '• 工作流调用接口实现'
            ],
            '交付成果': [
                '• 两个模板可用',
                '• 工作流集成完成',
                '• AI生成效果良好'
            ]
        }
    ]
    
    # 第三阶段：功能完善与上线（2周）
    phase3_tasks = [
        {
            '周次': '第5周',
            '核心任务': [
                '高级功能开发',
                '• 批量操作和任务队列',
                '• 协同编辑基础功能',
                '• 数据统计分析',
                '',
                '用户体验优化',
                '• UI界面美化',
                '• 响应式布局',
                '• 操作引导'
            ],
            '交付成果': [
                '• 批量功能可用',
                '• 界面焕然一新',
                '• 统计功能上线'
            ]
        },
        {
            '周次': '第6周',
            '核心任务': [
                '系统测试',
                '• 功能测试和bug修复',
                '• 性能优化和压力测试',
                '• 安全扫描',
                '',
                '上线准备',
                '• 用户文档编写',
                '• 培训材料准备',
                '• Docker环境封装',
                '• 正式发布上线'
            ],
            '交付成果': [
                '• 测试全部通过',
                '• 文档准备齐全',
                '• Docker镜像就绪',
                '• 系统正式上线'
            ]
        }
    ]
    
    # 组装数据
    phases = [
        ('第一阶段\n（系统安全与\nAI模型集成）', phase1_tasks),
        ('第二阶段\n（核心功能开发）', phase2_tasks),
        ('第三阶段\n（功能完善与上线）', phase3_tasks)
    ]
    
    for phase_name, tasks in phases:
        for i, task in enumerate(tasks):
            if i == 0:
                data['阶段'].append(phase_name)
            else:
                data['阶段'].append('')
            
            data['周次'].append(task['周次'])
            data['核心任务'].append('\n'.join(task['核心任务']))
            data['交付成果'].append('\n'.join(task['交付成果']))
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "项目实施计划"
    
    # 添加标题
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Word星辉点智能编辑器项目实施计划'
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    
    # 添加项目信息
    ws.merge_cells('A2:D2')
    ws['A2'] = '项目周期：6周 | 客户：国企培训部门 | 目标：正式上线运营'
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 35
    
    # 添加数据，从第4行开始
    start_row = 4
    
    # 添加表头
    headers = ['阶段', '周次', '核心任务', '交付成果']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置对齐方式
            if c_idx == 1:  # 阶段列
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if value:  # 如果有阶段名称，设置背景色
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                    cell.font = Font(bold=True, size=11)
            elif c_idx == 2:  # 周次列
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 合并阶段单元格
    current_phase = None
    phase_start_row = None
    
    for row in range(start_row + 1, ws.max_row + 1):
        phase_value = ws.cell(row=row, column=1).value
        if phase_value and phase_value != current_phase:
            if phase_start_row and current_phase:
                # 合并之前的阶段
                if row - 1 > phase_start_row:
                    ws.merge_cells(f'A{phase_start_row}:A{row-1}')
            current_phase = phase_value
            phase_start_row = row
    
    # 合并最后一个阶段
    if phase_start_row and phase_start_row < ws.max_row:
        ws.merge_cells(f'A{phase_start_row}:A{ws.max_row}')
    
    # 设置所有单元格边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # 调整行高
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[start_row].height = 30
    
    for row in range(start_row + 1, ws.max_row + 1):
        ws.row_dimensions[row].height = 80
    
    # 添加备注
    ws.merge_cells(f'A{ws.max_row + 2}:D{ws.max_row + 2}')
    note = '备注：核心功能包括AI模型集成、Word表格解析、MaxKB培训报告工作流，暂不包含用户认证系统。'
    ws[f'A{ws.max_row + 2}'] = note
    ws[f'A{ws.max_row + 2}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{ws.max_row + 2}'].font = Font(italic=True, size=10, color="666666")
    
    # 保存文件
    filename = f'Word_Clean_Implementation_Plan_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(filename)
    print(f"Clean implementation plan generated: {filename}")
    return filename

if __name__ == "__main__":
    create_clean_implementation_plan()